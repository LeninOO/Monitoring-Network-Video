
#!/usr/bin/env python

# NO debe aparecer  ningun ASI

import os
from socket  import * 
import threading
import SocketServer
from datetime import datetime, date, time, timedelta
import calendar

#from jsonrpclib import Server
import urllib
import  socket
import time
import sys
import time
from time import time, sleep
import commands
import math
from datetime import date
import xlwt
traduce={'Aug':'8'}

from Tkinter import *
master = Tk()
e = Entry(master)
e.pack()
piedra =[]




def convierte(DICCI,mezcal):
                aux3 = DICCI.get(mezcal)
                return aux3
def flujos(bloque):
     calidos=[' ']
     limite = len(bloque) 
     lim = limite -  3
     cursor1 =0
     while( cursor1 < lim ):
       temporal1 = bloque[cursor1]
       cadena1 = temporal1['stream']
       
       repetidos = 0
       for indice in calidos: 
            if(indice==cadena1):   
                repetidos = repetidos + 1       
       if(repetidos<=0):
                calidos.append(cadena1)  
       cursor1 = cursor1+1 
     #print "Flujos que tienen error :"
     #print calidos          
     return calidos          

def  estadisticas(bloque2,calido2):
      GENERAL ={}
      CORONEL ={}
      INDICES ={}
      guarda = []
      grande =[]
      total = 0
      errores = 0
      veces = 0
      acumulado = 0
      limite = len(bloque2) 
      lim = limite -  3
      cursor =0
      print limite
      
      for u in calido2 :
           
           cursor=0
           while( cursor < lim ):
                   tempo = bloque2[cursor]
                   nombre = tempo['stream']
                   tiempo = tempo['duration']
                   vector = tempo['index']
                   if(u==nombre):   
                          veces = veces + 1 
                          acumulado =acumulado + int(tiempo)
                          guarda.append(tempo)                                       
                   cursor = cursor+1
           CORONEL.setdefault(u,acumulado)
           GENERAL.setdefault(u,veces)
           INDICES.setdefault(u,guarda)
           temporal5 = acumulado
	   total = temporal5+total
	   temporal6 = veces
           errores= temporal6 + errores
           
           veces =0
           acumulado=0
           guarda = []  
      
      
      CORONEL.setdefault('TIEMPO_ALARMA',total)
      GENERAL.setdefault('ERRORES_FLUJO',errores)
      
      #print "-------------------------------------------------------------"
      #print GENERAL
      #print CORONEL
      grande.append(GENERAL)
      grande.append(CORONEL)
      grande.append(INDICES)      
      return grande
      

def  procesa( victima):

     ahora = datetime.now()  # Obtiene fecha y hora actual
     
     publica = " "
     dia= ahora.day
     
     mes= ahora.month
     
     tag = victima
     ruta =""+tag+".xml"
     ficha1 = open(ruta,'r' )
     lineas1=ficha1.readlines()
     limite1 = len(lineas1)
     lim = int(limite1)
     lim = lim - 2
     cursor1 = 2

     punterolinea= 0
     cabecera = 2
     general = lineas1[cabecera]
     calidos=['index','status','time','duration','type','stream','description','color']
     LIBRO = {}
     foco =[]
     soquete=[]
     punto = 0
     while( cursor1 < lim ):
       temporal1 = lineas1[cursor1]
       cadena1 = str(temporal1)
       inlinea = cadena1.split('"')
       
       for  q in calidos:
              if(str(q))=='index' :
                          valor = 3
              if(str(q))=='status' :
                          valor = 5
              if(str(q))=='time' :
                          valor = 7
              if(str(q))=='duration' :
                          valor = 9
              if(str(q))=='type' :
                          valor = 11
              if(str(q))=='stream' :
                          valor = 13
              if(str(q))=='description' :
                          valor = 15
                          
              if(str(q))=='color' :
                          valor = 17
                          
              LIBRO[q]= inlinea[valor]
       MESA =LIBRO 
       
       aux1 = LIBRO.get('time')
       aux2 = LIBRO.get('stream')
       if(len(aux2)>0):
              canal = aux2.split()
       else:
              canal = 'TNT HD'
       flujo = canal[0]
       
       fecha = aux1.split()
       mesito = fecha[0]
       me = convierte(traduce,str(mesito))
       dita = fecha[1]
       
              
       LIBRO = { }
       cursor1 = cursor1+1 
       punterolinea=0
       semaforo = False
       
       if(  (str(dita)==str(dia) ) and  ( str(flujo)!='ASI:') ):
                     semaforo = True 
       if(  (semaforo) and  (str(me)==str(mes) ) ):
                     foco.append(MESA)
      
       semaforo= False
        
     return  foco
     print len(foco)
     
     


def pastel(big,cali):
     
     from openpyxl import Workbook
     from openpyxl.chart import (
     PieChart3D,
     Reference
     )
     
     wb = Workbook()
     
     ws = wb.active
     ws1= wb.active
     ws2= wb.active
     
     ws = wb.create_sheet(0)
     ws1 = wb.create_sheet()
     ws2 = wb.create_sheet()
     
     ws.title =" DATOS"
     ws1.title = "ESTADISTICAS"
     ws2.title = "TIEMPOS"
     ws.sheet_properties.tabColor = "1072BA"
     ws1.sheet_properties.tabColor = "2012BA"
     ws2.sheet_properties.tabColor = "10021A"
     #-------------------------------------------------------------#
     flu=big[0]
     vagon=[]
     lista =[]
     for u in flu:
              lista.append(u)
              cazuela = flu.get(u)
              lista.append(cazuela)
              vagon.append(lista)
              lista=[]
     for u in vagon:
           ws.append(u)
     #-------------------------------------------------------------#
     flu1=big[1]
     vagon1=[]
     lista1 =[]
     for u1 in flu1:
              lista1.append(u1)
              cazuela1 = flu1.get(u1)
              lista1.append(cazuela1)
              vagon1.append(lista1)
              lista1=[]
     for u1 in vagon1:
           ws1.append(u1)
     #-------------------------------------------------------------#
     ca=['stream','status','time','duration','type','index','description','color']
     
     flu2=big[2]
     vagon2=[]
     vagon2.append(ca)
     lista2 =[]
     for u2 in cali:
               aux4 = flu2.get(u2)
               for v in aux4:
                      for w in ca :
                           aux5= v.get(w)
                           lista2.append(aux5)
                      vagon2.append(lista2)
                      lista2=[]
               
     
     for u4 in vagon2:
            ws2.append(u4)
     
     #-------------------------------------------------------------#
     
     
     w = len(cali)
     pie = PieChart3D()
     torta = PieChart3D()
     
     labels = Reference(ws, min_col=1, min_row=2, max_row=w)
     data = Reference(ws, min_col=2, min_row=1, max_row=w)
     pie.add_data(data, titles_from_data=True)
     pie.set_categories(labels)
     pie.title = "Distribucion de Alertas "
     ws.add_chart(pie, "D2")
     #-----------------------------------------------------------------#
     labels1 = Reference(ws1, min_col=1, min_row=2, max_row=w)
     data1 = Reference(ws1, min_col=2, min_row=1, max_row=w)
     torta.add_data(data1, titles_from_data=True)
     torta.set_categories(labels)
     torta.title = "Duracion de alertas por canal"
     ws1.add_chart(torta, "D2")
     labels2 = Reference(ws2, min_col=1, min_row=2, max_row=w)
     data2 = Reference(ws2, min_col=2, min_row=1, max_row=w)
     
     
     wb.save("fin.xlsx")
     
     
     
def callback():
           bandera = True
           esta = e.get()
           piedra = procesa(esta)
           cereza=  flujos(piedra)
           vela= estadisticas(piedra,cereza)
	   #libro(piedra,cereza)
	   pastel(vela,cereza)
	   #grafica()
   
   
b = Button(master, text="get", width=10, command=callback)
b.pack()

mainloop()


                                                     

	

   